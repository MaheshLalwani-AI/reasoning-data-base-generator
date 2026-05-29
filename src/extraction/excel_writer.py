from pathlib import Path

import openpyxl
from openpyxl.styles import Font, PatternFill
from openpyxl.utils import get_column_letter

from models import Question


class ExcelWriter:
    QUESTION_HEADERS = [
        "Exam Name",
        "Page Number",
        "Question Number",
        "Question",
        "Option 1",
        "Option 2",
        "Option 3",
        "Option 4",
        "Correct Answer",
        "AI Answer",
        "Verification Status",
        "Reasoning Type",
        "Regex Topic",
        "LLM Topic",
        "Solution",
    ]

    REPORT_HEADERS = [
        "PDF Name",
        "Pages Count",
        "Parsed Questions",
        "Detected Answer Keys",
        "Applied Answers",
        "Missing Answers",
        "Answer Coverage %",
        "Reasoning Questions Kept",
        "Error",
    ]

    def write(
        self,
        questions: list[Question],
        output_path: str,
        report_rows: list[dict] | None = None,
    ) -> None:
        output_file = Path(output_path)

        output_file.parent.mkdir(
            parents=True,
            exist_ok=True,
        )

        workbook = openpyxl.Workbook()

        questions_sheet = workbook.active
        questions_sheet.title = "Questions"

        self._write_questions_sheet(
            sheet=questions_sheet,
            questions=questions,
        )

        report_sheet = workbook.create_sheet(
            "Extraction Report"
        )

        self._write_report_sheet(
            sheet=report_sheet,
            report_rows=report_rows or [],
        )

        workbook.save(output_file)

    def _write_questions_sheet(
        self,
        sheet,
        questions: list[Question],
    ) -> None:
        sheet.append(self.QUESTION_HEADERS)
        self._style_header_row(sheet)

        for question in questions:
            options = list(question.options)

            while len(options) < 4:
                options.append("")

            row = [
                question.exam_name,
                question.page_number,
                question.question_number,
                question.question_text,
                options[0],
                options[1],
                options[2],
                options[3],
                question.correct_answer,
                question.ai_answer,
                question.verification_status,
                question.reasoning_type,
                question.regex_topic,
                question.llm_topic,
                question.solution,
            ]

            sheet.append(row)

        self._auto_width(sheet)

    def _write_report_sheet(
        self,
        sheet,
        report_rows: list[dict],
    ) -> None:
        sheet.append(self.REPORT_HEADERS)
        self._style_header_row(sheet)

        for report in report_rows:
            row = [
                report.get("pdf_name", ""),
                report.get("pages_count", 0),
                report.get("parsed_questions", 0),
                report.get("detected_answer_keys", 0),
                report.get("applied_answers", 0),
                report.get("missing_answers", 0),
                report.get("answer_coverage_percent", 0.0),
                report.get("reasoning_questions_kept", 0),
                report.get("error", ""),
            ]

            sheet.append(row)

        self._auto_width(sheet)

    def _style_header_row(self, sheet) -> None:
        fill = PatternFill(
            start_color="D9EAF7",
            end_color="D9EAF7",
            fill_type="solid",
        )

        for cell in sheet[1]:
            cell.font = Font(bold=True)
            cell.fill = fill

    def _auto_width(self, sheet) -> None:
        for column_cells in sheet.columns:
            max_length = 0
            column_letter = get_column_letter(
                column_cells[0].column
            )

            for cell in column_cells:
                value = cell.value

                if value is None:
                    continue

                max_length = max(
                    max_length,
                    len(str(value)),
                )

            sheet.column_dimensions[
                column_letter
            ].width = min(
                max(max_length + 2, 14),
                45,
            )